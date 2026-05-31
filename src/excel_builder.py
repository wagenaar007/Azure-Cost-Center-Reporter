import logging
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

C_DARK_BLUE  = "1E3A5F"
C_MED_BLUE   = "2E75B6"
C_LIGHT_BLUE = "EBF3FB"
C_TOTAL      = "BDD7EE"
C_WHITE      = "FFFFFF"
C_TEXT       = "1A1A2E"
C_GRAY       = "F5F7FA"

FILL_HEADER  = PatternFill("solid", fgColor=C_DARK_BLUE)
FILL_ALT     = PatternFill("solid", fgColor=C_LIGHT_BLUE)
FILL_TOTAL   = PatternFill("solid", fgColor=C_TOTAL)
FILL_GRAY    = PatternFill("solid", fgColor=C_GRAY)

FONT_TITLE   = Font(bold=True, size=16, color=C_DARK_BLUE, name="Calibri")
FONT_SUBTITLE= Font(bold=True, size=12, color=C_MED_BLUE,  name="Calibri")
FONT_HEADER  = Font(bold=True, size=10, color=C_WHITE,     name="Calibri")
FONT_TOTAL   = Font(bold=True, size=10, color=C_DARK_BLUE, name="Calibri")
FONT_NORMAL  = Font(size=10, color=C_TEXT, name="Calibri")
FONT_SMALL   = Font(size=9,  color="666666", name="Calibri")

ALIGN_C  = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L  = Alignment(horizontal="left",   vertical="center")
ALIGN_R  = Alignment(horizontal="right",  vertical="center")

_THIN = Side(style="thin", color="C5D8ED")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

NUM_FMT  = '#,##0.00 [$€-407]'
DATE_FMT = 'DD.MM.YYYY'

TAB_COLORS = {
    "Übersicht":     "1E3A5F",
    "Analyse":        "C0392B",
    "Dienste":       "E67E22",
    "Wöchentlich":   "3A8FC1",
    "Monatlich":     "4DA8D4",
    "Jährlich":      "70C1E8",
    "Ressourcen":    "27AE60",
    "Subscriptions": "8E44AD",
}


def _cell(ws: Worksheet, row: int, col: int, value: Any = None,
          font=None, fill=None, align=None, border=None, num_fmt: str = None):
    c = ws.cell(row=row, column=col, value=value)
    if font:    c.font      = font
    if fill:    c.fill      = fill
    if align:   c.alignment = align
    if border:  c.border    = border
    if num_fmt: c.number_format = num_fmt
    return c


def _header_row(ws: Worksheet, row: int, headers: list[str],
                col_widths: list[int] | None = None, start_col: int = 1):
    for i, h in enumerate(headers):
        col = start_col + i
        _cell(ws, row, col, h,
              font=FONT_HEADER, fill=FILL_HEADER, align=ALIGN_C, border=BORDER)
        ws.row_dimensions[row].height = 30
    if col_widths:
        for i, w in enumerate(col_widths):
            ws.column_dimensions[get_column_letter(start_col + i)].width = w


def _data_rows(ws: Worksheet, start_row: int, rows: list[list[Any]],
               cost_cols: list[int] | None = None,
               int_cols:  list[int] | None = None,
               date_cols: list[int] | None = None):
    for row_idx, row_data in enumerate(rows):
        r = start_row + row_idx
        fill = FILL_ALT if row_idx % 2 == 1 else None
        for col_idx, val in enumerate(row_data, 1):
            fmt = None
            aln = ALIGN_L
            if cost_cols and col_idx in cost_cols:
                fmt = NUM_FMT
                aln = ALIGN_R
            elif int_cols and col_idx in int_cols:
                aln = ALIGN_R
            elif date_cols and col_idx in date_cols:
                fmt = DATE_FMT
            _cell(ws, r, col_idx, val,
                  font=FONT_NORMAL, fill=fill, align=aln,
                  border=BORDER, num_fmt=fmt)
        ws.row_dimensions[r].height = 18


def _total_row(ws: Worksheet, row: int, values: list[Any],
               cost_cols: list[int] | None = None):
    for col_idx, val in enumerate(values, 1):
        fmt = NUM_FMT if cost_cols and col_idx in cost_cols else None
        _cell(ws, row, col_idx, val,
              font=FONT_TOTAL, fill=FILL_TOTAL, align=ALIGN_R if fmt else ALIGN_L,
              border=BORDER, num_fmt=fmt)
    ws.row_dimensions[row].height = 20


def _set_tab(ws: Worksheet, name: str):
    ws.sheet_properties.tabColor = TAB_COLORS.get(name, C_DARK_BLUE)


def _autofilter(ws: Worksheet, header_row: int, last_col: int):
    ws.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(last_col)}{header_row}"
    )
    ws.freeze_panes = f"A{header_row + 1}"


def _build_uebersicht(wb: Workbook, sub_totals: list[dict],
                      resource_totals: list[dict], date_from: str, date_to: str):
    ws = wb.create_sheet("Übersicht")
    _set_tab(ws, "Übersicht")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2

    ws.merge_cells("B2:H2")
    _cell(ws, 2, 2, "Azure Cost Center Report",
          font=FONT_TITLE, align=ALIGN_L)
    ws.row_dimensions[2].height = 36

    ws.merge_cells("B3:H3")
    period_text = f"Zeitraum: {date_from}  –  {date_to}   |   Erstellt: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    _cell(ws, 3, 2, period_text, font=FONT_SMALL, align=ALIGN_L)

    ws.merge_cells("B5:E5")
    _cell(ws, 5, 2, "Gesamtkosten je Subscription",
          font=FONT_SUBTITLE, align=ALIGN_L)

    headers = ["Subscription", "Gesamtkosten", "Währung", "Ressourcen"]
    widths  = [28, 28, 14, 10]
    _header_row(ws, 6, headers, widths, start_col=2)
    ws.column_dimensions["F"].hidden = True

    grand_total = 0.0
    chart_data_start = 7

    for i, s in enumerate(sub_totals):
        r = 7 + i
        row = [s["SubscriptionName"], s["TotalCost"], s.get("Currency", "EUR"), s["ResourceCount"]]
        fill = FILL_ALT if i % 2 == 1 else None
        for c_idx, val in enumerate(row, 2):
            fmt  = NUM_FMT if c_idx == 3 else None
            aln  = ALIGN_R if c_idx in (3, 5) else ALIGN_L
            _cell(ws, r, c_idx, val,
                  font=FONT_NORMAL, fill=fill,
                  align=aln, border=BORDER, num_fmt=fmt)
        ws.row_dimensions[r].height = 18
        grand_total += s["TotalCost"]

    total_row = 7 + len(sub_totals)
    _cell(ws, total_row, 2, "GESAMT", font=FONT_TOTAL, fill=FILL_TOTAL, align=ALIGN_L, border=BORDER)
    _cell(ws, total_row, 3, round(grand_total, 2),
          font=FONT_TOTAL, fill=FILL_TOTAL, align=ALIGN_R, border=BORDER, num_fmt=NUM_FMT)
    _cell(ws, total_row, 4, sub_totals[0].get("Currency", "EUR") if sub_totals else "EUR",
          font=FONT_TOTAL, fill=FILL_TOTAL, align=ALIGN_L, border=BORDER)
    _cell(ws, total_row, 5, sum(s["ResourceCount"] for s in sub_totals),
          font=FONT_TOTAL, fill=FILL_TOTAL, align=ALIGN_R, border=BORDER)
    ws.row_dimensions[total_row].height = 22

    top10_start_col = 8
    top10_start_row = 6
    ws.merge_cells(
        start_row=top10_start_row - 1, start_column=top10_start_col,
        end_row=top10_start_row - 1,   end_column=top10_start_col + 3
    )
    _cell(ws, top10_start_row - 1, top10_start_col,
          "Top 10 – teuerste Ressourcen (Gesamt)",
          font=FONT_SUBTITLE, align=ALIGN_L)

    top_hdrs  = ["Ressource", "Service", "Subscription", "Gesamtkosten"]
    top_widths = [28, 32, 22, 20]
    _header_row(ws, top10_start_row, top_hdrs, top_widths, start_col=top10_start_col)

    top10 = resource_totals[:10]
    for i, r in enumerate(top10):
        row_num = top10_start_row + 1 + i
        fill    = FILL_ALT if i % 2 == 1 else None
        vals    = [r["ResourceName"], r["ServiceName"],
                   r["SubscriptionName"], r["TotalCost"]]
        for c_idx, val in enumerate(vals, top10_start_col):
            fmt = NUM_FMT if c_idx == top10_start_col + 3 else None
            _cell(ws, row_num, c_idx, val,
                  font=FONT_NORMAL, fill=fill,
                  align=ALIGN_R if fmt else ALIGN_L,
                  border=BORDER, num_fmt=fmt)
        ws.row_dimensions[row_num].height = 18


def _build_weekly(wb: Workbook, records: list[dict]):
    ws = wb.create_sheet("Wöchentlich")
    _set_tab(ws, "Wöchentlich")

    headers = ["Woche", "Subscription", "Ressource", "Ressourcentyp",
               "Ressourcegruppe", "Service", "Kosten", "Währung"]
    widths  = [12, 22, 36, 28, 24, 22, 12, 9]
    _header_row(ws, 1, headers, widths)
    _autofilter(ws, 1, len(headers))

    rows = [
        [
            r.get("Period", ""),
            r.get("SubscriptionName", ""),
            r.get("ResourceName", ""),
            r.get("ResourceType", ""),
            r.get("ResourceGroup", ""),
            r.get("ServiceName", ""),
            float(r.get("Cost", 0)),
            r.get("Currency", "EUR"),
        ]
        for r in records
    ]
    _data_rows(ws, 2, rows, cost_cols=[7])

    total = sum(float(r.get("Cost", 0)) for r in records)
    _total_row(ws, len(rows) + 2,
               ["GESAMT", "", "", "", "", "", round(total, 2), ""],
               cost_cols=[7])


def _build_monthly(wb: Workbook, records: list[dict]):
    ws = wb.create_sheet("Monatlich")
    _set_tab(ws, "Monatlich")

    headers = ["Monat", "Subscription", "Ressource", "Ressourcentyp",
               "Ressourcegruppe", "Service", "Kosten", "Währung"]
    widths  = [12, 22, 36, 28, 24, 22, 12, 9]
    _header_row(ws, 1, headers, widths)
    _autofilter(ws, 1, len(headers))

    rows = [
        [
            r.get("Period", ""),
            r.get("SubscriptionName", ""),
            r.get("ResourceName", ""),
            r.get("ResourceType", ""),
            r.get("ResourceGroup", ""),
            r.get("ServiceName", ""),
            float(r.get("Cost", 0)),
            r.get("Currency", "EUR"),
        ]
        for r in records
    ]
    _data_rows(ws, 2, rows, cost_cols=[7])

    total = sum(float(r.get("Cost", 0)) for r in records)
    _total_row(ws, len(rows) + 2,
               ["GESAMT", "", "", "", "", "", round(total, 2), ""],
               cost_cols=[7])


def _build_yearly(wb: Workbook, records: list[dict]):
    ws = wb.create_sheet("Jährlich")
    _set_tab(ws, "Jährlich")

    headers = ["Jahr", "Subscription", "Ressource", "Ressourcentyp",
               "Ressourcegruppe", "Service", "Kosten", "Währung"]
    widths  = [8, 22, 36, 28, 24, 22, 12, 9]
    _header_row(ws, 1, headers, widths)
    _autofilter(ws, 1, len(headers))

    rows = [
        [
            r.get("Period", ""),
            r.get("SubscriptionName", ""),
            r.get("ResourceName", ""),
            r.get("ResourceType", ""),
            r.get("ResourceGroup", ""),
            r.get("ServiceName", ""),
            float(r.get("Cost", 0)),
            r.get("Currency", "EUR"),
        ]
        for r in records
    ]
    _data_rows(ws, 2, rows, cost_cols=[7])

    total = sum(float(r.get("Cost", 0)) for r in records)
    _total_row(ws, len(rows) + 2,
               ["GESAMT", "", "", "", "", "", round(total, 2), ""],
               cost_cols=[7])


def _build_resources(wb: Workbook, resource_totals: list[dict]):
    ws = wb.create_sheet("Ressourcen")
    _set_tab(ws, "Ressourcen")

    headers = ["Ressource", "Ressourcentyp", "Ressourcegruppe",
               "Service", "Subscription", "Gesamtkosten", "Währung"]
    widths  = [36, 28, 24, 22, 22, 14, 9]
    _header_row(ws, 1, headers, widths)
    _autofilter(ws, 1, len(headers))

    rows = [
        [
            r["ResourceName"],
            r["ResourceType"],
            r["ResourceGroup"],
            r["ServiceName"],
            r["SubscriptionName"],
            r["TotalCost"],
            r.get("Currency", "EUR"),
        ]
        for r in resource_totals
    ]
    _data_rows(ws, 2, rows, cost_cols=[6])

    total = sum(r["TotalCost"] for r in resource_totals)
    _total_row(ws, len(rows) + 2,
               ["GESAMT", "", "", "", "", round(total, 2), ""],
               cost_cols=[6])


def _build_charts(
    wb: Workbook,
    monthly_records: list[dict],
    resource_totals: list[dict],
    sub_totals: list[dict],
):
    from collections import defaultdict

    ws = wb.create_sheet("Analyse")
    _set_tab(ws, "Analyse")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2

    ws.merge_cells("B2:N2")
    _cell(ws, 2, 2, "Kostenanalyse", font=FONT_TITLE, align=ALIGN_L)
    ws.row_dimensions[2].height = 36

    all_months = sorted({r.get("Period", "") for r in monthly_records if r.get("Period")})
    subs       = [s["SubscriptionName"] for s in sub_totals]

    pivot: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for r in monthly_records:
        pivot[r.get("SubscriptionName", "")][r.get("Period", "")] += float(r.get("Cost", 0))

    ws.merge_cells("B4:D4")
    _cell(ws, 4, 2, "Monatliche Kostenübersicht je Subscription", font=FONT_SUBTITLE, align=ALIGN_L)

    piv_hdr    = ["Monat"] + subs + ["Gesamt", "Ggü. Vormonat"]
    piv_widths = [12] + [22] * len(subs) + [14, 13]
    _header_row(ws, 5, piv_hdr, piv_widths, start_col=2)
    last_piv_col = 1 + len(piv_hdr)
    CHANGE_COL   = last_piv_col
    ws.auto_filter.ref = f"B5:{get_column_letter(last_piv_col)}5"
    ws.freeze_panes = "B6"

    month_totals = [sum(pivot[sub].get(m, 0) for sub in subs) for m in all_months]

    grand_total = 0.0
    for i, month in enumerate(all_months):
        r_num = 6 + i
        fill  = FILL_ALT if i % 2 == 1 else None
        row_total = 0.0
        _cell(ws, r_num, 2, month, font=FONT_NORMAL, fill=fill, align=ALIGN_L, border=BORDER)
        for j, sub in enumerate(subs):
            val = round(pivot[sub].get(month, 0), 2)
            row_total += val
            _cell(ws, r_num, 3 + j, val,
                  font=FONT_NORMAL, fill=fill, align=ALIGN_R, border=BORDER, num_fmt=NUM_FMT)
        _cell(ws, r_num, 3 + len(subs), round(row_total, 2),
              font=FONT_NORMAL, fill=fill, align=ALIGN_R, border=BORDER, num_fmt=NUM_FMT)
        if i == 0 or month_totals[i - 1] == 0:
            chg_val  = "\u2013"
            chg_font = FONT_NORMAL
        else:
            pct = (month_totals[i] - month_totals[i - 1]) / month_totals[i - 1] * 100
            chg_val  = f"{'+' if pct >= 0 else ''}{pct:.1f} %"
            chg_font = Font(size=10, color="C0392B" if pct > 0 else "27AE60", name="Calibri")
        _cell(ws, r_num, CHANGE_COL, chg_val, font=chg_font, fill=fill, align=ALIGN_R, border=BORDER)
        grand_total += row_total
        ws.row_dimensions[r_num].height = 18

    total_row = 6 + len(all_months)
    _cell(ws, total_row, 2, "GESAMT", font=FONT_TOTAL, fill=FILL_TOTAL, align=ALIGN_L, border=BORDER)
    for j, sub in enumerate(subs):
        _cell(ws, total_row, 3 + j, round(sum(pivot[sub].values()), 2),
              font=FONT_TOTAL, fill=FILL_TOTAL, align=ALIGN_R, border=BORDER, num_fmt=NUM_FMT)
    _cell(ws, total_row, 3 + len(subs), round(grand_total, 2),
          font=FONT_TOTAL, fill=FILL_TOTAL, align=ALIGN_R, border=BORDER, num_fmt=NUM_FMT)
    _cell(ws, total_row, CHANGE_COL, "\u2013", font=FONT_NORMAL, fill=FILL_TOTAL, align=ALIGN_R, border=BORDER)
    ws.row_dimensions[total_row].height = 22

    svc_totals: dict[str, float] = defaultdict(float)
    for r in resource_totals:
        svc_totals[r.get("ServiceName", "Sonstige")] += r.get("TotalCost", 0)
    top_svcs = sorted(svc_totals.items(), key=lambda x: -x[1])[:15]

    SVC_COL = last_piv_col + 2
    ws.merge_cells(
        start_row=4, start_column=SVC_COL,
        end_row=4,   end_column=SVC_COL + 1
    )
    _cell(ws, 4, SVC_COL, "Top 15 Azure-Services", font=FONT_SUBTITLE, align=ALIGN_L)
    _header_row(ws, 5, ["Service", "Kosten (EUR)"], [32, 16], start_col=SVC_COL)
    for i, (svc, cost) in enumerate(top_svcs):
        fill = FILL_ALT if i % 2 == 1 else None
        _cell(ws, 6 + i, SVC_COL,     svc,            font=FONT_NORMAL, fill=fill, align=ALIGN_L, border=BORDER)
        _cell(ws, 6 + i, SVC_COL + 1, round(cost, 2), font=FONT_NORMAL, fill=fill, align=ALIGN_R, border=BORDER, num_fmt=NUM_FMT)
        ws.row_dimensions[6 + i].height = 18


def _build_service_monthly(wb: Workbook, monthly_records: list[dict]):
    from collections import defaultdict

    ws = wb.create_sheet("Dienste")
    _set_tab(ws, "Dienste")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2

    ws.merge_cells("B2:J2")
    _cell(ws, 2, 2, "Monatliche Kostenentwicklung je Azure-Service",
          font=FONT_TITLE, align=ALIGN_L)
    ws.row_dimensions[2].height = 36

    ws.merge_cells("B3:J3")
    _cell(ws, 3, 2,
          "Δ abs. / Δ % zeigen die Veränderung des letzten Monats gegenüber dem Vormonat.",
          font=FONT_SMALL, align=ALIGN_L)

    pivot: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for r in monthly_records:
        svc = r.get("ServiceName") or "Sonstige"
        m   = r.get("Period", "")
        if m:
            pivot[svc][m] += float(r.get("Cost", 0))

    all_months = sorted({r.get("Period", "") for r in monthly_records if r.get("Period")})
    n          = len(all_months)

    svc_totals = {svc: sum(costs.values()) for svc, costs in pivot.items()}
    all_svcs   = sorted(svc_totals, key=lambda s: -svc_totals[s])

    has_delta = n >= 2
    m_last    = all_months[-1] if has_delta else None
    m_prev    = all_months[-2] if has_delta else None

    total_col     = 3 + n
    delta_abs_col = 4 + n
    delta_pct_col = 5 + n
    last_col      = delta_pct_col if has_delta else total_col

    hdrs   = ["Service"] + all_months + ["Gesamt"]
    widths = [34] + [13] * n + [14]
    if has_delta:
        hdrs   += ["Δ gg. Vormonat (€)", "Δ %"]
        widths += [16, 11]
    _header_row(ws, 5, hdrs, widths, start_col=2)
    ws.auto_filter.ref = f"B5:{get_column_letter(last_col)}5"
    ws.freeze_panes    = "C6"

    month_totals: dict[str, float] = defaultdict(float)

    for i, svc in enumerate(all_svcs):
        r_num = 6 + i
        fill  = FILL_ALT if i % 2 == 1 else None

        _cell(ws, r_num, 2, svc, font=FONT_NORMAL, fill=fill, align=ALIGN_L, border=BORDER)

        row_total = 0.0
        for j, m in enumerate(all_months):
            val = round(pivot[svc].get(m, 0), 4)
            row_total       += val
            month_totals[m] += val
            _cell(ws, r_num, 3 + j, val,
                  font=FONT_NORMAL, fill=fill, align=ALIGN_R, border=BORDER, num_fmt=NUM_FMT)

        _cell(ws, r_num, total_col, round(row_total, 2),
              font=FONT_NORMAL, fill=fill, align=ALIGN_R, border=BORDER, num_fmt=NUM_FMT)

        if has_delta:
            v_last = pivot[svc].get(m_last, 0)
            v_prev = pivot[svc].get(m_prev, 0)
            d_abs  = round(v_last - v_prev, 2)
            d_font = Font(size=10, name="Calibri",
                          color=("C0392B" if d_abs > 0 else ("27AE60" if d_abs < 0 else C_TEXT)))
            _cell(ws, r_num, delta_abs_col, d_abs,
                  font=d_font, fill=fill, align=ALIGN_R, border=BORDER, num_fmt=NUM_FMT)

            if v_prev != 0:
                pct      = round((v_last - v_prev) / v_prev * 100, 1)
                pct_text = f"{'+' if pct >= 0 else ''}{pct:.1f} %"
                pct_font = Font(size=10, name="Calibri",
                                color=("C0392B" if pct > 0 else "27AE60"),
                                bold=abs(pct) >= 20)
            elif v_last == 0:
                pct_text, pct_font = "–", FONT_NORMAL
            else:
                pct_text = "neu"
                pct_font = Font(size=10, name="Calibri", color="8E44AD")
            _cell(ws, r_num, delta_pct_col, pct_text,
                  font=pct_font, fill=fill, align=ALIGN_R, border=BORDER)

        ws.row_dimensions[r_num].height = 18

    total_r = 6 + len(all_svcs)
    _cell(ws, total_r, 2, "GESAMT", font=FONT_TOTAL, fill=FILL_TOTAL, align=ALIGN_L, border=BORDER)
    grand = 0.0
    for j, m in enumerate(all_months):
        mt = round(month_totals[m], 2)
        grand += mt
        _cell(ws, total_r, 3 + j, mt,
              font=FONT_TOTAL, fill=FILL_TOTAL, align=ALIGN_R, border=BORDER, num_fmt=NUM_FMT)
    _cell(ws, total_r, total_col, round(grand, 2),
          font=FONT_TOTAL, fill=FILL_TOTAL, align=ALIGN_R, border=BORDER, num_fmt=NUM_FMT)
    if has_delta:
        t_last = month_totals.get(m_last, 0)
        t_prev = month_totals.get(m_prev, 0)
        td     = round(t_last - t_prev, 2)
        _cell(ws, total_r, delta_abs_col, td,
              font=FONT_TOTAL, fill=FILL_TOTAL, align=ALIGN_R, border=BORDER, num_fmt=NUM_FMT)
        if t_prev != 0:
            tp = round((t_last - t_prev) / t_prev * 100, 1)
            _cell(ws, total_r, delta_pct_col,
                  f"{'+' if tp >= 0 else ''}{tp:.1f} %",
                  font=FONT_TOTAL, fill=FILL_TOTAL, align=ALIGN_R, border=BORDER)
        else:
            _cell(ws, total_r, delta_pct_col, "–",
                  font=FONT_TOTAL, fill=FILL_TOTAL, align=ALIGN_R, border=BORDER)
    ws.row_dimensions[total_r].height = 22


def _build_subscriptions(wb: Workbook, sub_totals: list[dict]):
    ws = wb.create_sheet("Subscriptions")
    _set_tab(ws, "Subscriptions")

    headers = ["Subscription ID", "Name", "Gesamtkosten", "Währung", "Ressourcen"]
    widths  = [38, 28, 16, 10, 12]
    _header_row(ws, 1, headers, widths)

    rows = [
        [
            s["SubscriptionId"],
            s["SubscriptionName"],
            s["TotalCost"],
            s.get("Currency", "EUR"),
            s["ResourceCount"],
        ]
        for s in sub_totals
    ]
    _data_rows(ws, 2, rows, cost_cols=[3])

    total = sum(s["TotalCost"] for s in sub_totals)
    _total_row(ws, len(rows) + 2,
               ["", "GESAMT", round(total, 2), "", sum(s["ResourceCount"] for s in sub_totals)],
               cost_cols=[3])


def build_excel(
    output_file: str,
    daily_records:    list[dict],
    weekly_records:   list[dict],
    monthly_records:  list[dict],
    yearly_records:   list[dict],
    resource_totals:  list[dict],
    sub_totals:       list[dict],
    date_from:        str,
    date_to:          str,
) -> None:
    logger.info("Erstelle Excel-Workbook...")
    wb = Workbook()
    wb.remove(wb.active)

    _build_uebersicht      (wb, sub_totals, resource_totals, date_from, date_to)
    _build_charts          (wb, monthly_records, resource_totals, sub_totals)
    _build_service_monthly (wb, monthly_records)
    _build_weekly          (wb, weekly_records)
    _build_monthly         (wb, monthly_records)
    _build_yearly          (wb, yearly_records)
    _build_resources       (wb, resource_totals)
    _build_subscriptions   (wb, sub_totals)

    wb.save(output_file)
    logger.info(f"Excel gespeichert: {output_file}")
